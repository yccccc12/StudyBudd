import streamlit as st
import openpyxl
import requests
import json
import pandas as pd
from datetime import datetime
import re
from google.oauth2 import service_account
from googleapiclient.discovery import build
import googlemaps
import google.generativeai as genai
import folium
from streamlit_folium import folium_static
import polyline
from streamlit_js_eval import streamlit_js_eval


FILE_PATH = "StudyPlanner.xlsx"

# Set the page layout to wide    
st.set_page_config(page_title="StudyBudd", layout="wide")

# 📌 Function to call Gemini API
def load_gemini_api_key():
    try:
        with open("credentials.json", "r") as file:
            credentials = json.load(file)
            return credentials.get("gemini_api_key", None)
    except (FileNotFoundError, json.JSONDecodeError):
        return None  # Return None if file is missing or corrupted

def load_calendar_id():
    try:
        with open("credentials.json", "r") as file:
            calendar_data = json.load(file)
            return calendar_data.get("calendar_id", None)
    except (FileNotFoundError, json.JSONDecodeError):
        return None  # Return None if file is missing or corrupted

def load_google_maps_api_key():
    try:
        with open("credentials.json", "r") as file:
            maps_data = json.load(file)
            return maps_data.get("google_map_api_key", None)
    except (FileNotFoundError, json.JSONDecodeError):
        return None  # Return None if file is missing or corrupted
   
def ask_gemini_api_key(input_text):
    api_key = load_gemini_api_key()
    url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash:generateContent?key={api_key}"
    
    data = {
        "contents": {
            "role": "USER",
            "parts": [{"text": input_text}]
        },
        "generation_config": {
            "temperature": 0.3,
            "topP": 1,
            "maxOutputTokens": 256
        }
    }

    headers = {"Content-Type": "application/json"}
    
    response = requests.post(url, headers=headers, data=json.dumps(data))
    
    if response.status_code == 200:
        json_response = response.json()
        return json_response["candidates"][0]["content"]["parts"][0]["text"]
    else:
        st.error("Error: API request failed")
        return None

# 📌 Function to extract study details
def extract_study_details(user_input):
    prompt = f"""
    Summarize this study plan in plain text (no Markdown, no formatting). Provide details in this format:
    
    Event Name: [event]
    Date: [date]
    Time start: [time_start]
    Time end: [time_end]
    Time: [time]
    Priority: [priority]
    Notes: [notes]

    If only the time start is provided, then set the time end to one hour after the time start.
    Input: "{user_input}"
    """

    ai_response = ask_gemini_api_key(prompt)
    if not ai_response:
        return None

    lines = ai_response.split("\n")

    event_info = {
        "event_name": "",
        "date": "",
        "time_start": "",
        "time_end": "",
        "priority": "Medium",
        "notes": "No additional notes"
    }

    for line in lines:
        if ":" in line:
            key, value = line.split(":", 1)
            event_info[key.strip().lower().replace(" ", "_")] = value.strip()

    # Format the date properly
    event_info = get_formatted_date(event_info)

    return event_info

# 🎯 Function to extract date in YYYY-MM-DD format from AI's response
def extract_date_from_response(response):
    # Match the date pattern YYYY-MM-DD
    match = re.search(r"\d{4}-\d{2}-\d{2}", response)
    return match.group(0) if match else None

def get_formatted_date(event_info):
    if event_info.get("date"):
        # Get today's date in YYYY-MM-DD format
        today = datetime.today().strftime("%Y-%m-%d")
        
        # Create the prompt for AI model
        date_prompt = f"Today is {today}. What is the exact date for '{event_info['date']}'? Respond with only the date in YYYY-MM-DD format."
        
        # Get AI's response
        ai_response = ask_gemini_api_key(date_prompt)
        
        if ai_response:
            # Extract the date in YYYY-MM-DD format from AI's response
            match = extract_date_from_response(ai_response)
            if match:
                event_info["date"] = match
                print(f"Updated Event Info: {event_info}")
            else:
                print("AI response did not contain a valid date.")
        else:
            print("Failed to get a valid response from the AI.")
    return event_info


# 📌 Function to add the event to the Study_Plan sheet
def add_to_study_plan(event_name, event_date, event_time_start, event_time_end, priority, notes):
    wb = openpyxl.load_workbook(FILE_PATH)

    # Get the 'Study_Plan' sheet
    study_plan_sheet = wb["Study_Plan"]

    # Check if the sheet is empty or just has headers
    if study_plan_sheet.max_row == 1:  # Only header exists
        next_row = 2  # Start from the second row for data
        last_id_number = 1
    else:
        last_row = 1
        # Manually determine the next row by checking non-empty cells in the first column
        for row in range(study_plan_sheet.max_row, 1, -1):  # Loop backwards from max_row
            if study_plan_sheet.cell(row=row, column=1).value is not None:
                last_row = row
                break
        next_row = last_row + 1  # Append after the last non-empty row

        # Extract the last ID number and increment it
        last_id_cell = study_plan_sheet.cell(row=last_row, column=1).value
        if last_id_cell and last_id_cell.startswith("ID-"):
            last_id_number = int(last_id_cell.split('-')[1]) + 1  # Increment the last ID
        else:
            last_id_number = 1  # If no valid ID is found, start at 1

    # Generate the new ID
    new_id = f"ID-{last_id_number}"

    # Add the new entry to the sheet at the next available row
    study_plan_sheet.append([
        new_id,  # Insert the new ID at the start of the row
        event_name,
        event_date,
        event_time_start,
        event_time_end,
        priority,
        notes
    ])

    # Save the updated workbook
    wb.save(FILE_PATH)

# 📌 Function to display study plan in Streamlit
def display_study_plan():

    # Load Study_Plan sheet into Pandas DataFrame
    df = pd.read_excel(FILE_PATH, sheet_name="Study_Plan")

    if df.empty:
        st.warning("No study plan data available!")
    else:
        st.write("📚 **Your Study Plan**")
         # Add a checkbox column for selecting events
        df["Select"] = False  # Add a default column for selection

        # Use Streamlit's experimental data editor for direct editing
        # edited_df = st.data_editor(df, num_rows="dynamic", key="editable_table", use_container_width=True)
        edited_df = st.data_editor(df, num_rows="dynamic", key="editable_table")

        # Buttons for syncing events
        col1, col2 = st.columns(2)
        with col1:
            if st.button("Sync Selected Events"):
                # Filter selected rows
                selected_events = edited_df[edited_df["Select"] == True]
                if not selected_events.empty:
                    sync_with_google_calendar(selected_events)
                    st.success("✅ Selected events synced successfully!")
                else:
                    st.warning("⚠️ No events selected for syncing.")

        with col2:
            if st.button("Sync All Events"):
                sync_with_google_calendar(edited_df)
                st.success("✅ All events synced successfully!")
        
        # Save changes button
        if st.button("Save Changes"):
            # Save the edited DataFrame back to the Excel file
            edited_df.to_excel(FILE_PATH, sheet_name="Study_Plan", index=False)
            st.success("✅ Changes saved successfully!")
                # Display the calendar view below


# 📌 Function to sync events with Google Calendar
def sync_with_google_calendar(df):
    # Authenticate with Google Calendar API
    credentials = service_account.Credentials.from_service_account_file(
        "google_credentials.json",
        scopes=["https://www.googleapis.com/auth/calendar"]
    )
    calendarID = load_calendar_id()
    service = build("calendar", "v3", credentials=credentials)

    # Loop through the DataFrame and add events to the specified calendar
    for index, row in df.iterrows():
        try:
            # Validate and format the date and time
            event_date = pd.to_datetime(row["Date"]).strftime("%Y-%m-%d")
            
            # Prompt AI to convert the time to ISO 8601 format
            time_start_prompt = f"Convert the time '{row['Time Start']}' to ISO 8601 format. Respond with only the time in HH:MM:SS format."
            event_time_start = ask_gemini_api_key(time_start_prompt).replace("\n", "")

            # Prompt AI to convert the time to ISO 8601 format
            time_end_prompt = f"Convert the time '{row['Time End']}' to ISO 8601 format. Respond with only the time in HH:MM:SS format."
            event_time_end = ask_gemini_api_key(time_end_prompt).replace("\n", "")

            if not event_time_start or not event_time_end:
                raise ValueError(f"Missing time for event: {row['Event']}")
                
            # Construct the event object
            event = {
                "summary": row["Event"],
                "description": row.get("Notes", ""),  # Optional field
                "start": {
                    "dateTime": f"{event_date}T{event_time_start}",
                    "timeZone": "Asia/Kuala_Lumpur",  # Set to Malaysia timezone
                },
                "end": {
                    "dateTime": f"{event_date}T{event_time_end}",
                    "timeZone": "Asia/Kuala_Lumpur",  # Set to Malaysia timezone
                },
            }
            # Debug: Print the event data being sent
            print("Syncing event:", event)

            # Insert the event into Google Calendar
            service.events().insert(calendarId=calendarID, body=event).execute()
            print(f"Event synced successfully: {row['Event']}")

        except Exception as e:
            print(f"Details: {e}")

# 📌 Function to display the Google Calendar
def display_google_calendar():
    calendarID = load_calendar_id()
    st.subheader("📅 Google Calendar View")

    # Embed Google Calendar iframe
    calendar_url = f"https://calendar.google.com/calendar/embed?src={calendarID}&ctz=Asia/Kuala_Lumpur"
    st.components.v1.iframe(calendar_url, width=800, height=600)

# 📌 Streamlit UI
GOOGLE_MAPS_API_KEY = load_google_maps_api_key()
gmaps = googlemaps.Client(key=GOOGLE_MAPS_API_KEY)
genai.configure(api_key=load_gemini_api_key())

def find_nearest(model, user_location, use_current_location, user_input):
    #place_type = model.generate_content(f"Answer me in the form in either 'Primary School', 'Secondary School', 'University', 'School', or 'Library' if any form mentioned in '{user_input}', else answer 'Invalid'")
    #place type validation
    place_type = model.generate_content(f"is any education institution mentioned in '{user_input}'? (answer in form 'Primary School', 'Secondary School', 'University', 'School', or 'Library') else answer 'Invalid'")
    if 'Invalid' in place_type.text:
        st.error("Please enter valid types of education institution")
        return

    if 'No' in use_current_location.text:
            latitude, longitude = user_location.split(",")
    else:
        result = gmaps.geocode(f"{user_location.text}")
        latitude = result[0]["geometry"]["location"]["lat"]
        longitude = result[0]["geometry"]["location"]["lng"]

        
    if user_location:
        try:
            # Google Places API: Find nearest places
            places = gmaps.places_nearby(location= f"{latitude}, {longitude}", radius=5000, type=place_type.text.strip('\n').lower().replace(" ", "_"))
            
            if places['results']:
                st.subheader(f"Nearest {place_type.text.strip()}(s) :")
                m = folium.Map(location = [latitude, longitude], zoom_start=14)
                
                folium.TileLayer(
                    tiles='https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{z}/{y}/{x}',
                    attr='Esri',
                    name='Satellite',
                ).add_to(m)
                
                folium.TileLayer('OpenStreetMap').add_to(m)
                place_list = []
                
                for place in places['results']:  # Show top 5 places
                    if len(place_list) == 5:
                        break
                    
                    name = place['name']                   
                    is_valid = model.generate_content(f"Is '{name}' really is a {place_type.text.strip()}? (answer in form 'Yes' or 'No')")
                    
                    if 'No' in is_valid:
                        pass
                    else:
                        lat = place['geometry']['location']['lat']
                        lon = place['geometry']['location']['lng']
                        st.write(f"📍 {name}")
                        folium.Marker([lat, lon], tooltip=name, popup=name).add_to(m)      
                        place_list.append(place)                           

                # Add default OpenStreetMap layer (already has attribution)
                folium.LayerControl().add_to(m)

                folium_static(m)

                # Gemini AI: briefly introduce each places
                ai_response = model.generate_content(f"Introduce in detail among {', '.join([p['name'] for p in place_list])} in list form")
                st.subheader("Places Description")
                st.write(ai_response.text)

            else:
                st.error("❌ No nearby places found.")

        except Exception as e:
            st.error(f"❌ Error: {e}")
    
    
def find_route(model, user_location, user_input):
    #place_type = model.generate_content(f"Answer me in the form 'Valid' if any education institution exists in google map mentioned in '{user_input}', else answer 'Invalid'")
    #Validation to make sure user input education institution and not the others 
    place_type = model.generate_content(f"is any education institution exists in google map mentioned in '{user_input}'? (answer in form 'Valid' or 'Invalid')")
    if 'Invalid' in place_type.text:
        st.error("❌ Please enter valid types of education institution")
        return
    
    #Get destination from user input
    destination = model.generate_content(f"From '{user_input}', answer me the ending location only, no other word")
    directions = gmaps.directions(user_location.text, destination.text, mode="driving")
        
    if directions:
        route = directions[0]['overview_polyline']['points']
        decoded_route = polyline.decode(route)
        
        # Create a map centered on the route
        m = folium.Map(location=[decoded_route[0][0], decoded_route[0][1]], zoom_start=14)
        
        folium.TileLayer(
                    tiles='https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{z}/{y}/{x}',
                    attr='Esri',
                    name='Satellite',
                ).add_to(m)
                
        folium.TileLayer('OpenStreetMap').add_to(m)
        # Add the route to the map
        folium.PolyLine(
            locations=decoded_route,
            color='blue',
            weight=5
        ).add_to(m)
        
        folium.LayerControl().add_to(m)
        
        # Display distance and duration
        distance = directions[0]['legs'][0]['distance']['text']
        duration = directions[0]['legs'][0]['duration']['text']
        
        st.success(f"✅ Route found! Distance: {distance}, Duration: {duration}")
        folium_static(m)
    
    else:
        st.error("❌ No route found. Check your locations.")

# 📌 Streamlit UI
st.title("📖 AI-Powered Study Planner")

# Sidebar menu
menu = st.sidebar.selectbox("Choose an option", ["View Study Plan", "Update Study Plan", "View Calendar", "Generate Practice Questions", "Generate Flashcards", "Locate Educational Institution"])

if menu == "View Study Plan":
    display_study_plan()

elif menu == "Update Study Plan":
 # User input for event description
    st.subheader("Add a New Study Activity")

    # Text input for the user to describe their activity
    user_input = st.text_area("Describe your activity (e.g., 'I have a math test tomorrow at 5 pm')", "")

    # Button to process input and add to study plan
    if st.button("Schedule Study Activity"):
        if user_input:
            study_details = extract_study_details(user_input)
            if study_details:
                # Add to the study plan
                add_to_study_plan(
                    study_details["event_name"],
                    study_details["date"],
                    study_details["time_start"],
                    study_details["time_end"],
                    study_details["priority"],
                    study_details["notes"]
                )
                st.success("✅ Study activity scheduled successfully!")
                # Display updated study plan
                display_study_plan()
            else:
                st.error("❌ Failed to extract event details. Please try again.")
        else:
            st.error("❌ Please provide a description of your study activity.")

elif menu == "View Calendar":
    display_google_calendar()

elif menu == "Generate Practice Questions":
    st.subheader("Generate Practice Questions")

    # Text input for the user to provide a topic or prompt
    topic_prompt = st.text_area("Enter a topic or prompt for generating practice questions", "")

    # Button to generate practice questions
    if st.button("Generate Questions"):
        if topic_prompt:
            # Create a prompt for the AI to generate practice questions
            ai_prompt = f"""
            Generate 5 practice questions based on the following topic or prompt:
            "{topic_prompt}"
            Provide the questions in plain text format, numbered from 1 to 5.
            """
            questions = ask_gemini_api_key(ai_prompt)

            if questions:
                # Store questions in session state to persist across interactions
                st.session_state.questions = questions

                # Render LaTeX for mathematical equations if present
                st.write("### Practice Questions:")
                for line in questions.split("\n"):
                    if "$" in line:  # Check if the line contains LaTeX math symbols
                        st.latex(line.strip("$"))  # Render LaTeX
                    else:
                        st.write(line)  # Render as plain text
            else:
                st.error("❌ Failed to generate practice questions. Please try again.")
        else:
            st.error("❌ Please provide a topic or prompt.")

    # Button to show suggested solutions
    if "questions" in st.session_state and st.session_state.questions:
        if st.button("Show Suggested Solutions"):
            # Create a prompt for the AI to generate solutions
            solution_prompt = f"""
            Provide detailed solutions for the following practice questions:
            {st.session_state.questions}
            """
            solutions = ask_gemini_api_key(solution_prompt)

            if solutions:
                st.write("### Suggested Solutions:")
                for line in solutions.split("\n"):
                    if "$" in line:  # Check if the line contains LaTeX math symbols
                        st.latex(line.strip("$"))  # Render LaTeX
                    else:
                        st.write(line)  # Render as plain text
            else:
                st.error("❌ Failed to generate solutions. Please try again.")


elif menu == "Generate Flashcards":
    st.subheader("Generate Flashcards")

    # Text input for the user to provide a topic or content for flashcards
    flashcard_prompt = st.text_area("Enter a topic or content to generate flashcards", "")

    # Button to generate flashcards
    if st.button("Generate Flashcards"):
        if flashcard_prompt:
            # Create a prompt for the AI to generate flashcards
            ai_prompt = f"""
            Generate 10 flashcards based on the following topic or content:
            "{flashcard_prompt}"
            Provide the flashcards in the format:
            Question: [question]
            Answer: [answer]
            """
            flashcards = ask_gemini_api_key(ai_prompt)

            if flashcards:
                # Parse the flashcards into a structured format
                flashcard_list = []
                invalid_cards = []
                for card in flashcards.split("\n\n"):
                    if "Question:" in card and "Answer:" in card:
                        question = re.search(r"Question:\s*(.*)", card)
                        answer = re.search(r"Answer:\s*(.*)", card)
                        if question and answer:
                            flashcard_list.append({
                                "Question": question.group(1).strip(),
                                "Answer": answer.group(1).strip()
                            })
                        else:
                            invalid_cards.append(card)
                    else:
                        invalid_cards.append(card)

                # Display the flashcards
                if flashcard_list:
                    st.write("### Flashcards:")
                    for i, card in enumerate(flashcard_list, start=1):
                        with st.expander(f"Flashcard {i}: {card['Question']}"):
                            st.write(f"**Answer:** {card['Answer']}")
                else:
                    st.error("❌ Failed to parse flashcards. Or try to use topics which are suitable for flashcards. Please try again.")

            else:
                st.error("❌ Failed to generate flashcards. Please try again.")
        else:
            st.error("❌ Please provide a topic or content.")

elif menu == "Locate Educational Institution":
    st.title("🧭 AI-Powered Educational Institution Locator")
    st.write("Enter an educational institution related description, then AI will find it")
    
    model = genai.GenerativeModel("gemini-1.5-flash")
    user_input = st.text_input("Describe the educational institution related information to find nearest or route to it (etc. what are the nearest university from my location? \
                               What is the route from Rawang to university XXX ?)")
    #use_current_location = model.generate_content(f"answer me in the form 'Yes' or 'No', is '{user_input}' mention any starting location that exists in google map")
    #Validation for getting starting location, either mentioned or user own location by default
    use_current_location = model.generate_content(f"in '{user_input}', is there mentioned any location that exists in google map? (answer in form 'Yes' or 'No')")
    if st.button("Search") and user_input.strip() != "":
    
    #Obtaining user current location
        if 'No' in use_current_location.text:
            response = requests.get('https://ipinfo.io/json')
            result = response.json()
            user_location = result['loc']   
        else:
            #Getting user manual type in starting location
            user_location = model.generate_content(f"From '{user_input}', answer me the starting location only, no other word")
        
        #Mode validation to know what user want to find
        mode = model.generate_content(f"answer me in the form 'Find Nearest', 'Find Route', or 'Invalid', from '{user_input}'")
        
        if 'Find Nearest' in mode.text:
            find_nearest(model, user_location, use_current_location, user_input)
        elif 'Find Route' in mode.text:
            find_route(model, user_location, user_input)
        elif 'Invalid' in mode.text:
            st.error("❌ The input is unable to be process, please ask question related to finding nearest educational institution or route to it")
